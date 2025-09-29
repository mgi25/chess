# Swiss League Web App — Single‑file Flask app
import csv
import io
import math
import os
import re
from dataclasses import dataclass
from datetime import datetime
from typing import Dict, List, Optional, Tuple

from flask import Flask, request, redirect, url_for, render_template, send_file, flash
from sqlalchemy import (Column, Integer, String, Boolean, DateTime, ForeignKey,
                        create_engine, UniqueConstraint, func)
from sqlalchemy.orm import declarative_base, relationship, sessionmaker, scoped_session
import pandas as pd
import xlsxwriter
from jinja2 import DictLoader
from openpyxl import load_workbook

# - Auto standings with tiebreaks (Buchholz, Sonneborn‑Berger, wins)
# - Export snapshot to XLSX with multiple sheets (Players, Rounds, Standings)
# - Looks/behaves like a clean, grid‑first “Excel sheet” but in the browser
#
# Quick start
# 1) pip install -r requirements.txt
#    (Flask, SQLAlchemy, pandas, xlsxwriter)
# 2) python app.py
# 3) Open http://127.0.0.1:5000
#
# NOTE: This is a pragmatic Swiss pairing engine that satisfies common
# tournament needs. For FIDE‑strict edge cases you may need deeper constraints
# (forbidden pair lists, floaters across multiple rounds, fine‑grained color
# history rules, accelerated pairings, etc.). Those can be added in the
# `make_swiss_pairings()` function.
# -----------------------------------------------------------

from __future__ import annotations
import csv
import io
import math
import os
import statistics
from dataclasses import dataclass
from datetime import datetime
from typing import Dict, List, Optional, Tuple

from flask import Flask, request, redirect, url_for, render_template_string, send_file, flash
from sqlalchemy import (Column, Integer, String, Boolean, DateTime, ForeignKey,
                        create_engine, UniqueConstraint, func)
from sqlalchemy.orm import declarative_base, relationship, sessionmaker, scoped_session
import pandas as pd
import xlsxwriter

# ---------------------------------
# Flask / DB setup
# ---------------------------------
app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "dev-key")

DB_URL = os.environ.get("DATABASE_URL", "sqlite:///swiss.db")
engine = create_engine(DB_URL, echo=False, future=True)
SessionLocal = scoped_session(sessionmaker(bind=engine))
Base = declarative_base()


# ---------------------------------
# Models
# ---------------------------------
class Player(Base):
    __tablename__ = "players"
    id = Column(Integer, primary_key=True)
    name = Column(String, nullable=False)
    rating = Column(Integer, default=1200)
    club = Column(String, default="")
    bye_count = Column(Integer, default=0)

    def __repr__(self):
        return f"<Player {self.id} {self.name} ({self.rating})>"


class Round(Base):
    __tablename__ = "rounds"
    id = Column(Integer, primary_key=True)
    number = Column(Integer, nullable=False, unique=True)def tiebreaks(sess) -> Dict[int, Dict[str, float]]:
    """Calculate Buchholz and Sonneborn‑Berger for each player."""
    scores = get_scores(sess)
    om = opponents_map(sess)
    return {pid: {"buchholz": buchholz[pid], "sb": sb[pid]} for pid in scores}


def safe_rating(value, default: int = 1200) -> int:
    try:
        if value is None:
            return default
        if isinstance(value, (int, float)):
            if isinstance(value, float) and math.isnan(value):
                return default
            return int(round(float(value)))
        text = str(value).strip()
        if not text:
            return default
        return int(round(float(text)))
    except (ValueError, TypeError):
        return default


def parse_bool_cell(value) -> bool:
    if isinstance(value, bool):
        return value
    if value is None:
        return False
    if isinstance(value, (int, float)):
        return value != 0
    text = str(value).strip().lower()
    return text in {"1", "true", "yes", "y", "finished", "done", "x", "✓"}


def normalize_result(value) -> str:
    if value is None:
        return "*"
    text = str(value).strip()
    if not text:
        return "*"
    normalized = (
        text.lower()
        .replace(" ", "")
        .replace("–", "-")
        .replace("—", "-")
        .replace("½", "0.5")
    )
    mapping = {
        "1-0": "1-0",
        "1:0": "1-0",
        "0-1": "0-1",
        "0:1": "0-1",
        "0.5-0.5": "0.5-0.5",
        "0.5:0.5": "0.5-0.5",
        "1/2-1/2": "0.5-0.5",
        "1/2:1/2": "0.5-0.5",
        "bye": "BYE",
    }
    return mapping.get(normalized, "*")


def compute_standings(sess) -> List[Dict[str, float]]:
    scores = get_scores(sess)
    tb = tiebreaks(sess)
    wins: Dict[int, int] = {pid: 0 for pid in scores}
    for pr in sess.query(Pairing).filter(Pairing.finished == True).all():
        if pr.result == "1-0":
            wins[pr.white_id] = wins.get(pr.white_id, 0) + 1
        elif pr.result == "0-1":
            wins[pr.black_id] = wins.get(pr.black_id, 0) + 1
    rows = []
    for p in sess.query(Player).order_by(Player.id.asc()).all():
        rows.append({
            "pid": p.id,
            "name": p.name,
            "rating": p.rating,
            "score": scores.get(p.id, 0.0),
            "buchholz": tb.get(p.id, {}).get("buchholz", 0.0),
            "sb": tb.get(p.id, {}).get("sb", 0.0),
            "wins": wins.get(p.id, 0),
        })
    rows.sort(key=lambda r: (-r["score"], -r["buchholz"], -r["sb"], -r["wins"], -r["rating"], r["pid"]))
    return rows
@dataclass
class Seed:
    pid: int
    name: str
    rating: int
    score: float
    white_ct: int
    black_ct: int
    bye_count: int

    @property
    def color_balance(self) -> int:
        return self.white_ct - self.black_ct  # prefer closer to 0
    players = sess.query(Player).all()
    if not players:
        return []

    scores = get_scores(sess)
    colors = get_color_history(sess)
        s = Seed(
            pid=p.id,
            name=p.name,
            rating=p.rating or 1200,
            score=round(float(scores.get(p.id, 0.0)), 2),
            white_ct=colors.get(p.id, (0, 0))[0],
            black_ct=colors.get(p.id, (0, 0))[1],
            bye_count=p.bye_count,
        )
        seeds.append(s)
    if len(seeds) % 2 == 1:
        candidates = list(reversed(seeds))
        no_bye = [s for s in candidates if s.bye_count == 0]
        pool = no_bye if no_bye else candidates
        pool.sort(key=lambda s: (s.score, s.rating, s.pid))
        bye_pid = pool[0].pid
        # Remove the bye player from pairing pool
        seeds = [s for s in seeds if s.pid != bye_pid]

    round = relationship("Round", back_populates="pairings")
    white = relationship("Player", foreign_keys=[white_id])
    black = relationship("Player", foreign_keys=[black_id])

    __table_args__ = (
        UniqueConstraint("round_id", "board_no", name="uq_round_board"),
    )

    def __repr__(self):
        return f"<Pairing R{self.round_id} B{self.board_no} {self.white_id} vs {self.black_id} {self.result}>"


Base.metadata.create_all(engine)


# ---------------------------------
# Helpers: scoring, color history, opponents, tiebreaks
# ---------------------------------
Result = str  # alias

SCORE_MAP = {
    "1-0": (1.0, 0.0),
    "0-1": (0.0, 1.0),
    "0.5-0.5": (0.5, 0.5),
    "BYE": (1.0, 0.0),  # white gets a bye point
}


def get_all_rounds(sess) -> List[Round]:
    return sess.query(Round).order_by(Round.number.asc()).all()


def get_current_round_number(sess) -> int:
    r = sess.query(func.max(Round.number)).scalar()
    return int(r or 0)


def get_scores(sess) -> Dict[int, float]:
    """Compute total points per player from all finished results."""
    scores: Dict[int, float] = {p.id: 0.0 for p in sess.query(Player).all()}
    pairings: List[Pairing] = sess.query(Pairing).all()
    for pr in pairings:
        if pr.result in SCORE_MAP and pr.finished:
            w, b = pr.white_id, pr.black_id
            ws, bs = SCORE_MAP[pr.result]
            if pr.result == "BYE":
                # BYE is recorded as white_id gets 1 point, black_id is None
                scores[w] = scores.get(w, 0.0) + ws
            else:
                scores[w] = scores.get(w, 0.0) + ws
                scores[b] = scores.get(b, 0.0) + bs
    return scores


def get_color_history(sess) -> Dict[int, Tuple[int, int]]:
    """Return {player_id: (white_count, black_count)}"""
    hist = {p.id: (0, 0) for p in sess.query(Player).all()}
    for pr in sess.query(Pairing).all():
        if pr.white_id:
            w = hist[pr.white_id]
            hist[pr.white_id] = (w[0] + 1, w[1])
        if pr.black_id:
            b = hist[pr.black_id]
            hist[pr.black_id] = (b[0], b[1] + 1)
    return hist


def opponents_map(sess) -> Dict[int, set]:
    om: Dict[int, set] = {p.id: set() for p in sess.query(Player).all()}
    for pr in sess.query(Pairing).all():
        if pr.white_id and pr.black_id:
            om[pr.white_id].add(pr.black_id)
    rnd = Round(number=round_number)
BASE_HTML = """
<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Swiss League</title>
  <style>
    :root {
      color-scheme: light;
    }
    * { box-sizing: border-box; }
    body {
      font-family: 'Segoe UI', Tahoma, sans-serif;
      background: #f3f6fb;
      margin: 0;
      color: #1f2933;
    }
    a { color: inherit; }
    .page {
      max-width: 1100px;
      margin: 0 auto;
      padding: 32px 20px 48px;
    }
    header {
      display: flex;
      flex-wrap: wrap;
      align-items: center;
      justify-content: space-between;
      gap: 12px;
      margin-bottom: 24px;
    }
    nav {
      display: flex;
      gap: 8px;
      flex-wrap: wrap;
    }
    .card {
      background: #ffffff;
      border-radius: 14px;
      padding: 24px;
      box-shadow: 0 10px 28px rgba(15, 23, 42, 0.08);
    }
    .layout {
      display: grid;
      grid-template-columns: minmax(0, 2fr) minmax(0, 1fr);
      gap: 24px;
    }
    @media (max-width: 960px) {
      .layout {
        grid-template-columns: 1fr;
      }
    }
    .table {
      width: 100%;
      border-collapse: collapse;
      font-size: 14px;
    }
    .table th {
      text-align: left;
      padding: 10px;
      background: #eef2f7;
      font-size: 11px;
      letter-spacing: 0.08em;
      text-transform: uppercase;
      color: #475569;
    }
    .table td {
      padding: 10px;
      border-top: 1px solid #e2e8f0;
    }
    .btn {
      display: inline-flex;
      align-items: center;
      gap: 6px;
      padding: 8px 16px;
      border-radius: 8px;
      border: none;
      font-weight: 600;
      background: #1f2937;
      color: #ffffff;
      text-decoration: none;
      cursor: pointer;
      transition: background 0.2s ease;
    }
    .btn:hover { background: #111827; }
    .btn-secondary {
      background: #e2e8f0;
      color: #1f2937;
    }
    .btn-secondary:hover { background: #cbd5f5; }
    .btn-disabled, .btn:disabled {
      opacity: 0.55;
      cursor: not-allowed;
      background: #9ca3af;
    }
    .tag {
      display: inline-block;
      padding: 2px 10px;
      border-radius: 999px;
      font-size: 12px;
      background: #e2e8f0;
      color: #1f2937;
    }
    .flash {
      background: #fff7ed;
      border: 1px solid #fbc38d;
      color: #9a3412;
      padding: 12px 16px;
      border-radius: 10px;
      margin-bottom: 16px;
    }
    form.inline {
      display: inline-flex;
      gap: 8px;
      align-items: center;
      flex-wrap: wrap;
    }
    input[type="text"], input[type="number"], input[type="file"], select {
      padding: 8px 10px;
      border-radius: 8px;
      border: 1px solid #cbd5f5;
      font-size: 14px;
    }
    label.checkbox {
      font-size: 12px;
      display: inline-flex;
      align-items: center;
      gap: 4px;
    }
    footer {
      margin-top: 40px;
      font-size: 12px;
      color: #64748b;
      text-align: center;
    }
  </style>
</head>
<body>
  <div class="page">
    <header>
      <h1 style="margin:0;font-size:28px;font-weight:700;">Swiss League</h1>
      <nav>
        <a class="btn btn-secondary" href="{{ url_for('index') }}">Standings</a>
        <a class="btn btn-secondary" href="{{ url_for('players') }}">Players</a>
        <a class="btn btn-secondary" href="{{ url_for('rounds') }}">Rounds</a>
        <a class="btn btn-secondary" href="{{ url_for('import_xlsm') }}">Import XLSM</a>
        <a class="btn" href="{{ url_for('export_xlsx') }}">Export XLSX</a>
      </nav>
    </header>
    {% with messages = get_flashed_messages() %}
      {% if messages %}
        {% for m in messages %}
          <div class="flash">{{ m }}</div>
        {% endfor %}
      {% endif %}
    {% endwith %}
    {% block content %}{% endblock %}
    <footer>Built for Excel-style Swiss tournaments.</footer>
  </div>
</body>
</html>
"""

INDEX_HTML = """
{% extends 'base.html' %}
{% block content %}
<div class="layout">
  <section class="card">
    <div style="display:flex;justify-content:space-between;align-items:center;gap:12px;margin-bottom:16px;flex-wrap:wrap;">
      <h2 style="margin:0;font-size:20px;">Standings</h2>
      <form method="post" action="{{ url_for('generate_round') }}">
        <button class="btn {% if not can_generate %}btn-disabled{% endif %}" {% if not can_generate %}disabled{% endif %}>Generate Round {{ next_round }}</button>
      </form>
    </div>
    <div style="overflow-x:auto;">
      <table class="table">
        <thead>
          <tr>
            <th>#</th>
            <th>Player</th>
            <th>Rating</th>
            <th>Pts</th>
            <th>Buchholz</th>
            <th>Sonneborn-Berger</th>
            <th>Wins</th>
          </tr>
        </thead>
        <tbody>
          {% for row in table %}
          <tr>
            <td>{{ loop.index }}</td>
            <td style="font-weight:600;">{{ row.name }}</td>
            <td>{{ row.rating }}</td>
            <td style="font-weight:600;">{{ '%.1f'|format(row.score) }}</td>
            <td>{{ '%.1f'|format(row.buchholz) }}</td>
            <td>{{ '%.1f'|format(row.sb) }}</td>
            <td>{{ row.wins }}</td>
          </tr>
          {% endfor %}
        </tbody>
      </table>
    </div>
  </section>

  <section class="card">
    <h2 style="margin-top:0;font-size:20px;">Quick Actions</h2>
    <ul style="list-style:none;padding-left:0;margin:0;display:grid;gap:8px;">
      <li><a class="btn btn-secondary" href="{{ url_for('players') }}">Add/Import Players</a></li>
      <li><a class="btn btn-secondary" href="{{ url_for('rounds') }}">Enter Results</a></li>
      <li><a class="btn btn-secondary" href="{{ url_for('export_xlsx') }}">Export XLSX Snapshot</a></li>
    </ul>
    <p style="margin-top:18px;font-size:14px;color:#475569;line-height:1.4;">This mirrors your Excel flow: maintain Players → generate Round pairings → enter results → view real-time Standings → export.</p>
  </section>
</div>
{% endblock %}
"""

PLAYERS_HTML = """
{% extends 'base.html' %}
{% block content %}
<section class="card">
  <div style="display:flex;justify-content:space-between;gap:12px;flex-wrap:wrap;align-items:flex-end;margin-bottom:18px;">
    <h2 style="margin:0;font-size:20px;">Players</h2>
    <form method="post" class="inline">
      <input type="text" name="name" placeholder="Name" required>
      <input type="number" name="rating" placeholder="Rating" value="1200">
      <input type="text" name="club" placeholder="Club (optional)">
      <button class="btn" type="submit">Add</button>
    </form>
  </div>

  <div style="margin-bottom:24px;">
    <form method="post" action="{{ url_for('import_csv') }}" enctype="multipart/form-data" class="inline">
      <label style="font-weight:600;">Import CSV</label>
      <input type="file" name="file" accept=".csv" required>
      <button class="btn" type="submit">Upload</button>
      <span style="font-size:12px;color:#64748b;">Headers: name,rating,club</span>
    </form>
  </div>

  <div style="overflow-x:auto;">
    <table class="table">
      <thead>
        <tr>
          <th>ID</th>
          <th>Name</th>
          <th>Rating</th>
          <th>Club</th>
          <th>Byes</th>
          <th>Actions</th>
        </tr>
      </thead>
      <tbody>
        {% for p in players %}
        <tr>
          <td>{{ p.id }}</td>
          <td style="font-weight:600;">{{ p.name }}</td>
          <td>{{ p.rating }}</td>
          <td>{{ p.club }}</td>
          <td>{{ p.bye_count }}</td>
          <td>
            <a class="btn btn-secondary" href="{{ url_for('delete_player', pid=p.id) }}" onclick="return confirm('Delete player?')">Delete</a>
          </td>
        </tr>
        {% endfor %}
      </tbody>
    </table>
  </div>
</section>
{% endblock %}
"""

ROUNDS_HTML = """
{% extends 'base.html' %}
{% block content %}
<section class="card">
  <div style="display:flex;justify-content:space-between;gap:12px;flex-wrap:wrap;align-items:center;margin-bottom:20px;">
    <h2 style="margin:0;font-size:20px;">Rounds &amp; Pairings</h2>
    <form method="post" action="{{ url_for('generate_round') }}">
      <button class="btn">Generate Round {{ next_round }}</button>
    </form>
  </div>

  {% for r in rounds %}
  <div style="border:1px solid #e2e8f0;border-radius:12px;padding:16px;margin-bottom:16px;background:#fafbff;">
    <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:12px;flex-wrap:wrap;gap:10px;">
      <h3 style="margin:0;font-size:18px;">Round {{ r.number }}</h3>
      <span class="tag">{{ r.pairings|length }} boards</span>
    </div>
    <div style="overflow-x:auto;">
      <table class="table">
        <thead>
          <tr>
            <th>Board</th>
            <th>White</th>
            <th>Black</th>
            <th>Result</th>
            <th>Actions</th>
          </tr>
        </thead>
        <tbody>
          {% for pr in r.pairings|sort(attribute='board_no') %}
          <tr>
            <td>{{ pr.board_no }}</td>
            <td>{{ pr.white.name if pr.white else '-' }}</td>
            <td>{{ pr.black.name if pr.black else '-' }}</td>
            <td style="font-weight:600;">{{ pr.result }}</td>
            <td>
              {% if pr.black is none %}
                <span class="tag">BYE</span>
              {% else %}
              <form method="post" action="{{ url_for('update_result', pairing_id=pr.id) }}" class="inline">
                <select name="result">
                  <option value="*" {% if pr.result=='*' %}selected{% endif %}>*</option>
                  <option value="1-0" {% if pr.result=='1-0' %}selected{% endif %}>1-0</option>
                  <option value="0-1" {% if pr.result=='0-1' %}selected{% endif %}>0-1</option>
                  <option value="0.5-0.5" {% if pr.result=='0.5-0.5' %}selected{% endif %}>½-½</option>
                </select>
                <label class="checkbox"><input type="checkbox" name="started" value="1" {% if pr.started %}checked{% endif %}> started</label>
                <label class="checkbox"><input type="checkbox" name="finished" value="1" {% if pr.finished %}checked{% endif %}> finished</label>
                <button class="btn btn-secondary" type="submit">Save</button>
              </form>
              {% endif %}
            </td>
          </tr>
          {% endfor %}
        </tbody>
      </table>
    </div>
  </div>
  {% endfor %}
</section>
{% endblock %}
"""

IMPORT_XLSM_HTML = """
{% extends 'base.html' %}
{% block content %}
<section class="card" style="max-width:720px;margin:0 auto;">
  <h2 style="margin-top:0;font-size:20px;">Import SwissChessLeaguev4.xlsm</h2>
  <p style="font-size:14px;color:#475569;line-height:1.5;">Provide a file path or upload the workbook. Visible values are imported so the players, past rounds, and standings mirror Excel.</p>
  <form method="post" enctype="multipart/form-data" style="display:grid;gap:16px;margin-top:24px;">
    <label style="display:grid;gap:8px;font-size:14px;">
      <span style="font-weight:600;">Workbook path (optional)</span>
      <input type="text" name="path" placeholder="SwissChessLeaguev4.xlsm" value="{{ default_path }}">
    </label>
    <label style="display:grid;gap:8px;font-size:14px;">
      <span style="font-weight:600;">Or upload .xlsm</span>
      <input type="file" name="file" accept=".xlsm,.xlsx">
    </label>
    <div style="display:flex;justify-content:flex-end;gap:12px;flex-wrap:wrap;">
      <a class="btn btn-secondary" href="{{ url_for('index') }}">Cancel</a>
      <button class="btn" type="submit">Import Workbook</button>
    </div>
  </form>
</section>
{% endblock %}
"""

# Register templates
app.jinja_loader = DictLoader({
    "base.html": BASE_HTML,
    "index.html": INDEX_HTML,
    "players.html": PLAYERS_HTML,
    "rounds.html": ROUNDS_HTML,
    "import_xlsm.html": IMPORT_XLSM_HTML,
})
@app.route("/")
def index():
    sess = SessionLocal()
    try:
        rows = compute_standings(sess)
        cur = get_current_round_number(sess)
        can_gen = sess.query(Player).count() >= 2
        return render_template(
            "index.html",
            table=rows,
            can_generate=can_gen,
            next_round=cur + 1,
        )
    finally:
        sess.close()


@app.route("/players", methods=["GET", "POST"])
def players():
    sess = SessionLocal()
    try:
        if request.method == "POST":
            name = request.form.get("name", "").strip()
            rating = safe_rating(request.form.get("rating"), 1200)
            club = request.form.get("club", "").strip()
            if name:
                sess.add(Player(name=name, rating=rating, club=club))
                sess.commit()
                flash("Player added.")
            return redirect(url_for("players"))

        return render_template("players.html", players=sess.query(Player).order_by(Player.id.asc()).all())
    finally:
        sess.close()
        for row in rdr:
            name = (row.get("name") or "").strip()
            if not name:
                continue
            rating = safe_rating(row.get("rating"), 1200)
            club = (row.get("club") or "").strip()
            sess.add(Player(name=name, rating=rating, club=club))
            count += 1
@app.route("/rounds")
def rounds():
    sess = SessionLocal()
    try:
        return render_template(
            "rounds.html",
            rounds=get_all_rounds(sess),
            next_round=get_current_round_number(sess) + 1,
        )
    finally:
        sess.close()


@app.route("/generate_round", methods=["POST"])
def generate_round():
    sess = SessionLocal()
    try:
        if sess.query(Player).count() < 2:
            flash("Add at least two players to generate pairings.")
            return redirect(url_for("rounds"))
        next_round = get_current_round_number(sess) + 1
        pairings = make_swiss_pairings(sess, next_round)
        if pairings:
            flash(f"Generated {len(pairings)} pairings for Round {next_round}.")
        else:
            flash("No pairings generated.")
    finally:
        sess.close()
    return redirect(url_for("rounds"))


@app.route("/update_result/<int:pairing_id>", methods=["POST"])
def update_result(pairing_id: int):
    sess = SessionLocal()
    try:
        pr = sess.get(Pairing, pairing_id)
        flash("Result updated.")
    finally:
        sess.close()
    return redirect(url_for("rounds"))


@app.route("/import_xlsm", methods=["GET", "POST"])
def import_xlsm():
    default_path = os.path.abspath("SwissChessLeaguev4.xlsm")
    if request.method == "GET":
        return render_template("import_xlsm.html", default_path=default_path)

    file = request.files.get("file")
    path_value = (request.form.get("path") or "").strip()
    chosen_path = path_value or default_path

    try:
        if file and file.filename:
            data = file.read()
            if not data:
                raise ValueError("Uploaded file is empty.")
            stream = io.BytesIO(data)
            workbook = load_workbook(stream, data_only=True, keep_vba=True)
            source_desc = file.filename
        else:
            abs_path = os.path.abspath(chosen_path)
            if not os.path.exists(abs_path):
                flash(f"Workbook not found: {abs_path}")
                return redirect(url_for("import_xlsm"))
            workbook = load_workbook(abs_path, data_only=True, keep_vba=True)
            source_desc = abs_path
    except Exception as exc:
        flash(f"Import failed: {exc}")
        return redirect(url_for("import_xlsm"))

    def first_header(sheet):
        for idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            values = []
            for cell in row:
                if cell is None:
                    values.append("")
                else:
                    values.append(str(cell).strip())
            if not any(values):
                continue
            lowered = [v.lower() for v in values]
            mapping = {lowered[i]: i for i in range(len(lowered)) if lowered[i]}
            return idx, mapping
        return None, {}

    def find_col(mapping: Dict[str, int], *keys: str) -> Optional[int]:
        normalised = {k.replace(" ", ""): v for k, v in mapping.items()}
        for key in keys:
            key_norm = key.replace(" ", "").lower()
            for header, idx in normalised.items():
                if header == key_norm or header.endswith(key_norm):
                    return idx
                if key_norm in header:
                    return idx
        return None

    sess = SessionLocal()
    try:
        added_players = 0
        updated_players = 0
        rounds_imported = 0
        pairings_imported = 0

        existing_players = {p.name.strip().casefold(): p for p in sess.query(Player).all()}

        # Identify player sheet
        players_sheet = None
        for name in workbook.sheetnames:
            if name.strip().lower() == "players":
                players_sheet = workbook[name]
                break
        if players_sheet is None:
            for name in workbook.sheetnames:
                sheet = workbook[name]
                header_idx, mapping = first_header(sheet)
                if not mapping:
                    continue
                if "name" in mapping and ("rating" in mapping or "elo" in mapping):
                    players_sheet = sheet
                    break

        if players_sheet is not None:
            header_idx, mapping = first_header(players_sheet)
            name_idx = find_col(mapping, "name")
            rating_idx = find_col(mapping, "rating", "elo")
            club_idx = find_col(mapping, "club", "team")
            start_row = (header_idx or 0) + 1
            if name_idx is None:
                name_idx = 0
            for row in players_sheet.iter_rows(min_row=start_row, values_only=True):
                values = list(row)
                if not any(str(v).strip() if v is not None else "" for v in values):
                    continue
                name_val = values[name_idx] if name_idx is not None and name_idx < len(values) else None
                name = str(name_val).strip() if name_val is not None else ""
                if not name:
                    continue
                rating_val = values[rating_idx] if rating_idx is not None and rating_idx < len(values) else None
                club_val = values[club_idx] if club_idx is not None and club_idx < len(values) else ""
                rating = safe_rating(rating_val, 1200)
                club = str(club_val).strip() if club_val is not None else ""
                key = name.casefold()
                player = existing_players.get(key)
                if player:
                    if rating_val not in (None, ""):
                        player.rating = rating
                    if club:
                        player.club = club
                    updated_players += 1
                else:
                    player = Player(name=name, rating=rating, club=club)
                    sess.add(player)
                    sess.flush()
                    existing_players[key] = player
                    added_players += 1

        # Reset rounds and bye counters
        sess.query(Pairing).delete()
        sess.query(Round).delete()
        for player in sess.query(Player).all():
            player.bye_count = 0
        sess.flush()

        # Build round sheets
        round_sheets: List[Tuple[int, object]] = []
        for name in workbook.sheetnames:
            match = re.search(r"round\s*(\d+)", name, re.IGNORECASE)
            if match:
                round_number = int(match.group(1))
                round_sheets.append((round_number, workbook[name]))
        round_sheets.sort(key=lambda x: x[0])

        for round_no, sheet in round_sheets:
            header_idx, mapping = first_header(sheet)
            board_idx = find_col(mapping, "board")
            white_idx = find_col(mapping, "white")
            black_idx = find_col(mapping, "black")
            result_idx = find_col(mapping, "result")
            started_idx = find_col(mapping, "started")
            finished_idx = find_col(mapping, "finished")
            start_row = (header_idx or 0) + 1

            rnd = Round(number=round_no)
            sess.add(rnd)
            sess.flush()
            rounds_imported += 1

            board_counter = 1
            round_pairs = 0
            for row in sheet.iter_rows(min_row=start_row, values_only=True):
                values = list(row)
                if not any(str(v).strip() if v is not None else "" for v in values):
                    continue

                def take(idx: Optional[int]):
                    if idx is None or idx >= len(values):
                        return None
                    return values[idx]

                board_val = take(board_idx)
                try:
                    board_no = int(str(board_val).strip()) if board_val not in (None, "") else board_counter
                except (ValueError, TypeError):
                    board_no = board_counter
                white_name = str(take(white_idx)).strip() if take(white_idx) is not None else ""
                black_raw = take(black_idx)
                black_name = str(black_raw).strip() if black_raw is not None else ""
                result_raw = take(result_idx)
                normalized_result = normalize_result(result_raw)
                started = parse_bool_cell(take(started_idx))
                finished = parse_bool_cell(take(finished_idx))

                if not white_name and not black_name:
                    continue

                def ensure_player(name: str) -> Player:
                    key = name.casefold()
                    player = existing_players.get(key)
                    if player is None:
                        player = Player(name=name, rating=1200, club="")
                        sess.add(player)
                        sess.flush()
                        existing_players[key] = player
                    return player

                white_player = ensure_player(white_name) if white_name else None
                black_player = ensure_player(black_name) if black_name else None

                is_bye = normalized_result == "BYE" or (black_player is None and black_name == "")
                if is_bye and white_player is None:
                    continue
                if not is_bye and (white_player is None or black_player is None):
                    continue

                pairing = Pairing(round_id=rnd.id, board_no=board_no)
                pairing.white_id = white_player.id if white_player else None
                pairing.black_id = None if is_bye else (black_player.id if black_player else None)

                if is_bye:
                    pairing.result = "BYE"
                    pairing.started = True
                    pairing.finished = True
                    if white_player:
                        white_player.bye_count += 1
                else:
                    pairing.result = normalized_result
                    pairing.started = started or normalized_result in {"1-0", "0-1", "0.5-0.5"}
                    pairing.finished = finished or pairing.result in {"1-0", "0-1", "0.5-0.5"}

                sess.add(pairing)
                pairings_imported += 1
                round_pairs += 1
                board_counter = max(board_counter + 1, board_no + 1)

            if round_pairs == 0:
                sess.delete(rnd)
                rounds_imported -= 1

        sess.commit()
        flash(
            f"Imported {added_players} new players, updated {updated_players} players, "
            f"{rounds_imported} rounds, and {pairings_imported} pairings from {source_desc}."
        )
    finally:
        sess.close()

    return redirect(url_for("index"))


@app.route("/export.xlsx")
def export_xlsx():
    sess = SessionLocal()
    try:
        # Build DataFrames
        players = sess.query(Player).order_by(Player.id.asc()).all()
        standings_rows = []
        for row in compute_standings(sess):
            standings_rows.append({
                "Player ID": row["pid"],
                "Name": row["name"],
                "Rating": row["rating"],
                "Score": row["score"],
                "Buchholz": row["buchholz"],
                "Sonneborn-Berger": row["sb"],
            })
        standings_df = pd.DataFrame(standings_rows)

        rounds = sess.query(Round).order_by(Round.number.asc()).all()

        output = io.BytesIO()
        with pd.ExcelWriter(output, engine="xlsxwriter") as writer:
# Dev seed: add sample players if DB empty
# ---------------------------------
def seed_if_empty():
    sess = SessionLocal()
    try:
        if sess.query(Player).count() == 0:
            for i, (n, r) in enumerate([
                ("Alpha", 1800), ("Bravo", 1700), ("Charlie", 1650), ("Delta", 1600),
        sess.close()


if __name__ == "__main__":
    seed_if_empty()
    app.run(debug=True)
        if a.color_balance > b.color_balance:
            white_id, black_id = (b.pid, a.pid)
        pairings.append((white_id, black_id))

    # Insert bye as (pid, None)
    if bye_pid is not None:
        pairings.append((bye_pid, None))

    # Persist Round & Pairings
    nxt = get_current_round_number(sess) + 1
    rnd = Round(number=nxt)
    sess.add(rnd)
    sess.flush()

    # board numbers: 1..N, keep bye at last board
    for i, (w, b) in enumerate(pairings, 1):
        pr = Pairing(round_id=rnd.id, board_no=i, white_id=w, black_id=b)
        if b is None:
            pr.result = "BYE"
            pr.started = True
            pr.finished = True
            # increment player's bye_count
            pl = sess.get(Player, w)
            pl.bye_count += 1
        sess.add(pr)

    sess.commit()
    return pairings


# ---------------------------------
# UI templates (inline Jinja)
# ---------------------------------
BASE_HTML = """
<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Swiss League</title>
  <script src="https://cdn.tailwindcss.com"></script>
  <link rel="preconnect" href="https://fonts.googleapis.com">
  <link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
  <link href="https://fonts.googleapis.com/css2?family=Inter:wght@400;600;700&display=swap" rel="stylesheet">
  <style>
    body { font-family: Inter, ui-sans-serif, system-ui; }
    .card { @apply bg-white/90 dark:bg-slate-900/70 backdrop-blur rounded-2xl shadow p-6; }
    .gridhead { @apply text-xs uppercase tracking-wider text-slate-500; }
    .btn { @apply inline-flex items-center gap-2 rounded-xl px-3 py-2 text-sm font-semibold bg-slate-900 text-white hover:bg-slate-700; }
    .btn-sec { @apply inline-flex items-center gap-2 rounded-xl px-3 py-2 text-sm font-semibold bg-slate-200 hover:bg-slate-300 text-slate-800; }
    .tag { @apply text-xs px-2 py-0.5 rounded-full; }
  </style>
</head>
<body class="min-h-screen bg-gradient-to-br from-slate-50 to-slate-200/80 dark:from-slate-950 dark:to-slate-900 text-slate-900 dark:text-slate-100">
  <div class="max-w-7xl mx-auto p-6 space-y-6">
    <header class="flex items-center justify-between">
      <h1 class="text-2xl md:text-3xl font-bold">Swiss League</h1>
      <nav class="flex gap-2">
        <a class="btn-sec" href="{{ url_for('index') }}">Standings</a>
        <a class="btn-sec" href="{{ url_for('players') }}">Players</a>
        <a class="btn-sec" href="{{ url_for('rounds') }}">Rounds</a>
        <a class="btn" href="{{ url_for('export_xlsx') }}">Export XLSX</a>
      </nav>
    </header>
    {% with messages = get_flashed_messages() %}
      {% if messages %}
      <div class="space-y-2">
        {% for m in messages %}
          <div class="card border border-amber-300/60 bg-amber-50 text-amber-900">{{ m }}</div>
        {% endfor %}
      </div>
      {% endif %}
    {% endwith %}
    {% block content %}{% endblock %}
    <footer class="text-xs text-slate-500 py-6">Built for Excel‑style Swiss tournaments.</footer>
  </div>
</body>
</html>
"""

INDEX_HTML = """
{% extends 'base.html' %}
{% block content %}
<div class="grid grid-cols-1 lg:grid-cols-3 gap-6">
  <section class="lg:col-span-2 card">
    <div class="flex items-center justify-between mb-4">
      <h2 class="text-xl font-semibold">Standings</h2>
      <form method="post" action="{{ url_for('generate_round') }}">
        <button class="btn" {% if not can_generate %}disabled class="opacity-50"{% endif %}>Generate Round {{ next_round }}</button>
      </form>
    </div>
    <div class="overflow-x-auto">
      <table class="min-w-full text-sm">
        <thead class="gridhead">
          <tr class="text-left">
            <th class="p-2">#</th>
            <th class="p-2">Player</th>
            <th class="p-2">Rating</th>
            <th class="p-2">Pts</th>
            <th class="p-2">Buchholz</th>
            <th class="p-2">Sonneborn‑Berger</th>
            <th class="p-2">Wins</th>
          </tr>
        </thead>
        <tbody>
          {% for row in table %}
          <tr class="border-t border-slate-200/60">
            <td class="p-2">{{ loop.index }}</td>
            <td class="p-2 font-medium">{{ row.name }}</td>
            <td class="p-2">{{ row.rating }}</td>
            <td class="p-2 font-semibold">{{ '%.1f'|format(row.score) }}</td>
            <td class="p-2">{{ '%.1f'|format(row.buchholz) }}</td>
            <td class="p-2">{{ '%.1f'|format(row.sb) }}</td>
            <td class="p-2">{{ row.wins }}</td>
          </tr>
          {% endfor %}
        </tbody>
      </table>
    </div>
  </section>

  <section class="card">
    <h2 class="text-xl font-semibold mb-4">Quick Actions</h2>
    <ul class="space-y-2">
      <li><a class="btn-sec" href="{{ url_for('players') }}">Add/Import Players</a></li>
      <li><a class="btn-sec" href="{{ url_for('rounds') }}">Enter Results</a></li>
      <li><a class="btn-sec" href="{{ url_for('export_xlsx') }}">Export XLSX Snapshot</a></li>
    </ul>
    <div class="mt-6 text-sm text-slate-600 dark:text-slate-400">
      <p>This mirrors your Excel flow: maintain Players → generate Round pairings → enter results → view real‑time Standings → export.</p>
    </div>
  </section>
</div>
{% endblock %}
"""

PLAYERS_HTML = """
{% extends 'base.html' %}
{% block content %}
<section class="card">
  <div class="flex items-center justify-between mb-4">
    <h2 class="text-xl font-semibold">Players</h2>
    <form method="post" class="flex gap-2">
      <input type="text" name="name" placeholder="Name" class="px-3 py-2 rounded-xl border" required>
      <input type="number" name="rating" placeholder="Rating" class="px-3 py-2 rounded-xl border" value="1200">
      <input type="text" name="club" placeholder="Club (optional)" class="px-3 py-2 rounded-xl border">
      <button class="btn" type="submit">Add</button>
    </form>
  </div>

  <div class="mb-6">
    <form method="post" action="{{ url_for('import_csv') }}" enctype="multipart/form-data" class="flex items-center gap-3">
      <label class="font-semibold">Import CSV</label>
      <input type="file" name="file" accept=".csv" class="px-3 py-2 rounded-xl border" required>
      <button class="btn" type="submit">Upload</button>
      <span class="text-xs text-slate-500">Headers: name,rating,club</span>
    </form>
  </div>

  <div class="overflow-x-auto">
    <table class="min-w-full text-sm">
      <thead class="gridhead">
        <tr class="text-left">
          <th class="p-2">ID</th><th class="p-2">Name</th><th class="p-2">Rating</th><th class="p-2">Club</th><th class="p-2">Byes</th><th class="p-2">Actions</th>
        </tr>
      </thead>
      <tbody>
        {% for p in players %}
        <tr class="border-t border-slate-200/60">
          <td class="p-2">{{ p.id }}</td>
          <td class="p-2 font-medium">{{ p.name }}</td>
          <td class="p-2">{{ p.rating }}</td>
          <td class="p-2">{{ p.club }}</td>
          <td class="p-2">{{ p.bye_count }}</td>
          <td class="p-2">
            <a class="btn-sec" href="{{ url_for('delete_player', pid=p.id) }}" onclick="return confirm('Delete player?')">Delete</a>
          </td>
        </tr>
        {% endfor %}
      </tbody>
    </table>
  </div>
</section>
{% endblock %}
"""

ROUNDS_HTML = """
{% extends 'base.html' %}
{% block content %}
<section class="card space-y-6">
  <div class="flex items-center justify-between">
    <h2 class="text-xl font-semibold">Rounds & Pairings</h2>
    <form method="post" action="{{ url_for('generate_round') }}">
      <button class="btn">Generate Round {{ next_round }}</button>
    </form>
  </div>

  {% for r in rounds %}
  <div class="border rounded-2xl p-4">
    <div class="flex items-center justify-between mb-3">
      <h3 class="font-semibold">Round {{ r.number }}</h3>
      <span class="tag bg-slate-200 text-slate-800">{{ r.pairings|length }} boards</span>
    </div>
    <div class="overflow-x-auto">
      <table class="min-w-full text-sm">
        <thead class="gridhead">
          <tr class="text-left">
            <th class="p-2">Board</th>
            <th class="p-2">White</th>
            <th class="p-2">Black</th>
            <th class="p-2">Result</th>
            <th class="p-2">Actions</th>
          </tr>
        </thead>
        <tbody>
          {% for pr in r.pairings|sort(attribute='board_no') %}
          <tr class="border-t border-slate-200/60">
            <td class="p-2">{{ pr.board_no }}</td>
            <td class="p-2">{{ pr.white.name if pr.white else '-' }}</td>
            <td class="p-2">{{ pr.black.name if pr.black else '-' }}</td>
            <td class="p-2 font-medium">{{ pr.result }}</td>
            <td class="p-2">
              {% if pr.black is none %}
                <span class="tag bg-emerald-100 text-emerald-700">BYE</span>
              {% else %}
              <form method="post" action="{{ url_for('update_result', pairing_id=pr.id) }}" class="flex flex-wrap items-center gap-2">
                <select name="result" class="px-2 py-1 rounded border">
                  <option value="*" {% if pr.result=='*' %}selected{% endif %}>*</option>
                  <option value="1-0" {% if pr.result=='1-0' %}selected{% endif %}>1‑0</option>
                  <option value="0-1" {% if pr.result=='0-1' %}selected{% endif %}>0‑1</option>
                  <option value="0.5-0.5" {% if pr.result=='0.5-0.5' %}selected{% endif %}>½‑½</option>
                </select>
                <label class="inline-flex items-center gap-1 text-xs"><input type="checkbox" name="started" value="1" {% if pr.started %}checked{% endif %}> started</label>
                <label class="inline-flex items-center gap-1 text-xs"><input type="checkbox" name="finished" value="1" {% if pr.finished %}checked{% endif %}> finished</label>
                <button class="btn-sec" type="submit">Save</button>
              </form>
              {% endif %}
            </td>
          </tr>
          {% endfor %}
        </tbody>
      </table>
    </div>
  </div>
  {% endfor %}
</section>
{% endblock %}
"""

# Register templates
app.jinja_loader.mapping = {
    'base.html': BASE_HTML,
    'index.html': INDEX_HTML,
    'players.html': PLAYERS_HTML,
    'rounds.html': ROUNDS_HTML,
}


# ---------------------------------
# Routes
# ---------------------------------
@app.route("/")
def index():
    sess = SessionLocal()
    try:
        # standings
        scores = get_scores(sess)
        tb = tiebreaks(sess)

        # wins count
        win_count: Dict[int, int] = {pid: 0 for pid in scores}
        for pr in sess.query(Pairing).filter(Pairing.finished == True).all():
            if pr.result == "1-0":
                win_count[pr.white_id] = win_count.get(pr.white_id, 0) + 1
            elif pr.result == "0-1":
                win_count[pr.black_id] = win_count.get(pr.black_id, 0) + 1

        rows = []
        for p in sess.query(Player).all():
            rows.append({
                "pid": p.id,
                "name": p.name,
                "rating": p.rating,
                "score": scores.get(p.id, 0.0),
                "buchholz": tb.get(p.id, {}).get("buchholz", 0.0),
                "sb": tb.get(p.id, {}).get("sb", 0.0),
                "wins": win_count.get(p.id, 0)
            })

        rows.sort(key=lambda r: (-r["score"], -r["buchholz"], -r["sb"], -r["wins"], -r["rating"], r["pid"]))

        cur = get_current_round_number(sess)
        can_gen = sess.query(Player).count() >= 2
        return render_template_string(
            INDEX_HTML,
            table=rows,
            can_generate=can_gen,
            next_round=cur + 1
        )
    finally:
        sess.close()


@app.route("/players", methods=["GET", "POST"])
def players():
    sess = SessionLocal()
    try:
        if request.method == "POST":
            name = request.form.get("name", "").strip()
            rating = int(request.form.get("rating", 1200))
            club = request.form.get("club", "").strip()
            if name:
                sess.add(Player(name=name, rating=rating, club=club))
                sess.commit()
                flash("Player added.")
            return redirect(url_for("players"))

        return render_template_string(PLAYERS_HTML, players=sess.query(Player).order_by(Player.id.asc()).all())
    finally:
        sess.close()


@app.route("/players/delete/<int:pid>")
def delete_player(pid: int):
    sess = SessionLocal()
    try:
        p = sess.get(Player, pid)
        if p:
            # Prevent deletion if already paired
            used = sess.query(Pairing).filter((Pairing.white_id == pid) | (Pairing.black_id == pid)).first()
            if used:
                flash("Cannot delete: player already has pairings.")
            else:
                sess.delete(p)
                sess.commit()
                flash("Player deleted.")
        return redirect(url_for("players"))
    finally:
        sess.close()


@app.route("/import_csv", methods=["POST"])
def import_csv():
    file = request.files.get("file")
    if not file:
        flash("No file uploaded.")
        return redirect(url_for("players"))
    sess = SessionLocal()
    try:
        content = file.stream.read().decode("utf-8")
        rdr = csv.DictReader(io.StringIO(content))
        count = 0
        for row in rdr:
            name = (row.get("name") or "").strip()
            if not name:
                continue
            rating = int(row.get("rating") or 1200)
            club = (row.get("club") or "").strip()
            sess.add(Player(name=name, rating=rating, club=club))
            count += 1
        sess.commit()
        flash(f"Imported {count} players.")
    finally:
        sess.close()
    return redirect(url_for("players"))


@app.route("/rounds")
def rounds():
    sess = SessionLocal()
    try:
        return render_template_string(
            ROUNDS_HTML,
            rounds=get_all_rounds(sess),
            next_round=get_current_round_number(sess) + 1,
        )
    finally:
        sess.close()


@app.route("/generate_round", methods=["POST"]) 
def generate_round():
    sess = SessionLocal()
    try:
        pairings = make_swiss_pairings(sess, get_current_round_number(sess) + 1)
        flash(f"Generated {len(pairings)} pairings.")
    finally:
        sess.close()
    return redirect(url_for("rounds"))


@app.route("/update_result/<int:pairing_id>", methods=["POST"]) 
def update_result(pairing_id: int):
    sess = SessionLocal()
    try:
        pr = sess.get(Pairing, pairing_id)
        if not pr:
            flash("Pairing not found.")
            return redirect(url_for("rounds"))
        res = request.form.get("result", "*")
        started = request.form.get("started") == "1"
        finished = request.form.get("finished") == "1"
        if res not in {"*", "1-0", "0-1", "0.5-0.5"}:
            res = "*"
        pr.result = res if pr.black_id is not None else "BYE"
        pr.started = started or pr.result in {"1-0", "0-1", "0.5-0.5", "BYE"}
        pr.finished = finished or pr.result in {"1-0", "0-1", "0.5-0.5", "BYE"}
        sess.commit()
        flash("Result updated.")
    finally:
        sess.close()
    return redirect(url_for("rounds"))


@app.route("/export.xlsx")
def export_xlsx():
    sess = SessionLocal()
    try:
        # Build DataFrames
        players = sess.query(Player).order_by(Player.id.asc()).all()
        scores = get_scores(sess)
        tbs = tiebreaks(sess)

        standings_rows = []
        for p in players:
            standings_rows.append({
                "Player ID": p.id,
                "Name": p.name,
                "Rating": p.rating,
                "Score": scores.get(p.id, 0.0),
                "Buchholz": tbs.get(p.id, {}).get("buchholz", 0.0),
                "Sonneborn-Berger": tbs.get(p.id, {}).get("sb", 0.0),
            })
        standings_df = pd.DataFrame(standings_rows).sort_values(
            by=["Score", "Buchholz", "Sonneborn-Berger", "Rating", "Player ID"], ascending=[False, False, False, False, True]
        )

        rounds = sess.query(Round).order_by(Round.number.asc()).all()

        output = io.BytesIO()
        with pd.ExcelWriter(output, engine="xlsxwriter") as writer:
            # Players sheet
            pd.DataFrame([
                {"ID": p.id, "Name": p.name, "Rating": p.rating, "Club": p.club, "Byes": p.bye_count}
                for p in players
            ]).to_excel(writer, index=False, sheet_name="Players")

            # Round sheets
            for r in rounds:
                rows = []
                for pr in sorted(r.pairings, key=lambda x: x.board_no):
                    rows.append({
                        "Board": pr.board_no,
                        "White": pr.white.name if pr.white else "-",
                        "Black": pr.black.name if pr.black else "-",
                        "Result": pr.result,
                        "Started": pr.started,
                        "Finished": pr.finished,
                    })
                pd.DataFrame(rows).to_excel(writer, index=False, sheet_name=f"Round {r.number}")

            # Standings
            standings_df.to_excel(writer, index=False, sheet_name="Standings")

        output.seek(0)
        return send_file(output, as_attachment=True, download_name="SwissLeagueSnapshot.xlsx",
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    finally:
        sess.close()


# ---------------------------------
# Dev seed: add sample players if DB empty
# ---------------------------------
def maybe_seed():
    sess = SessionLocal()
    try:
        if sess.query(Player).count() == 0:
            for i, (n, r) in enumerate([
                ("Alpha", 1800), ("Bravo", 1700), ("Charlie", 1650), ("Delta", 1600),
                ("Echo", 1550), ("Foxtrot", 1500), ("Golf", 1450), ("Hotel", 1400),
            ], start=1):
                sess.add(Player(name=n, rating=r, club=""))
            sess.commit()
    finally:
        sess.close()


if __name__ == "__main__":
    maybe_seed()
    app.run(debug=True)
